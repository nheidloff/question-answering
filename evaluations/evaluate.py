#!/usr/bin/env python

# Imports
import requests
import json
import argparse
import os
import csv
from dotenv import load_dotenv
load_dotenv()
from datasets import load_metric
import evaluate as eval
import sacrebleu as scb
import logging
import openpyxl
import time
from datetime import date

# ******************************************
# Global variables

# qa service
endpoint = os.environ.get("endpoint")
api_url = os.environ.get("api_url")
username=os.environ.get("username")
password=os.environ.get("password")
verify_answer=os.environ.get("verify_answer")

# inputs
input_excel_filename = os.environ.get("input_excel_filename")
input_folder_name = os.environ.get("input_folder_name")
input_folder_name_qa_service_metrics = os.environ.get("input_folder_name_qa_service_metrics")

# outputs
output_question_resp_anwser = os.environ.get("output_question_resp_anwser")
output_question_resp_anwser_excel = os.environ.get("output_question_resp_anwser_excel")
output_error_log = os.environ.get("output_error_log")
output_session_id = os.environ.get("output_session_id")
output_folder_name = os.environ.get("output_folder_name")
number_of_retrys = os.environ.get("number_of_retrys")

print("*************************")
print("Environment configurations of 'evalute':")
print("")
print(f"- Endpoint: {endpoint}")
print(f"- API URL: {api_url}")
print(f"- Retrys: {number_of_retrys}")
print(f"- Username: {username}")
print(f"- Password: {password}")
print(f"- Verify answer: {verify_answer}\n")
print(f"- Input Excel: {input_excel_filename}")
print(f"- Input folder name: {input_folder_name}")
print(f"- Input folder name qa service: {input_folder_name_qa_service_metrics}\n")
print(f"- Output folder name: {output_folder_name}")
print(f"- Sesssion ID output prefix: {output_session_id}")
print(f"- Output CSV: {output_question_resp_anwser}")
print(f"- Output Excel: {output_question_resp_anwser_excel}\n")
print(f"- Error log name: {output_error_log}")

# ******************************************
# get os path information
def get_input_path():
        global input_folder_name
        directory = os.getcwd()
        directory = directory + "/" + input_folder_name
        return directory

def get_output_path():
        global output_folder_name
        directory = os.getcwd()
        directory = directory + "/" + output_folder_name
        return directory

def get_input_qa_service_metrics_local_path():
        global input_folder_name_qa_service_metrics
        directory = os.getcwd()
        directory = directory + "/../metrics/" + input_folder_name_qa_service_metrics
        return directory

def get_input_qa_service_metrics_container_path():
        global input_folder_name_qa_service_metrics
        directory = os.getcwd()
        directory = directory + "/metrics/"
        return directory

# ******************************************
# Bleu prepare eval data functions 

def bleu_run(input_filename):
    header, rows = bleu_get_data(input_filename)
    header = list(header.keys())
    return header, rows

def bleu_get_data(input_filename):
    rows = list()
    header = ['question', 'response', 'golden_anwser']

    try:
        workbook = openpyxl.load_workbook(input_filename)
        ws = workbook['experiment_data']
    except:
        print(f"Error: Could not open {input_filename}\n")
    
    for row in ws.iter_rows(values_only=True):
        rows.append(row)

    return bleu_from_list_to_dict(header), rows

def bleu_from_list_to_dict(header):
    indices = [i for i in range(0, len(header))]
    header = {k: v for k, v in zip(header, indices)}
    # print ( header )
    return header

# ******************************************
# Define logging
logger = logging.getLogger(str(date.today()) + "_" + output_session_id + "_" + output_error_log)
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler(get_output_path() + "/" + str(date.today()) + "_" + output_session_id + "_" + output_error_log)
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# Select the active worksheet
def create_output_workbook (workbook_name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet("experiment_data")
        worksheet.title = "experiment_data"

        worksheet['A1'] = 'question'
        worksheet['B1'] = 'answer'
        worksheet['C1'] = 'golden_anwser'
        worksheet['D1'] = 'golden_passage_1'
        worksheet['E1'] = 'golden_passage_2'
        worksheet['F1'] = 'golden_passage_3'
        worksheet['G1'] = 'answer_passage_1'
        worksheet['H1'] = 'answer_passage_2'
        worksheet['I1'] = 'answer_passage_3'

        # Add a header to the worksheet
        # header_text = 'Temp Response Values'
        # worksheet.header_footer.oddHeader.text = header_text

        # Save the workbook as a new Excel file
        workbook.save(workbook_name)
        return workbook

# ******************************************
# load input excel
def load_input_excel(excel_input):
    wb = openpyxl.load_workbook(excel_input)
    ws = wb.active
    
    rows = []
    for rdx, row in enumerate(ws.iter_rows(values_only=True)):
        if rdx:
            rows.append(list(row))
        else:
            header = row
    
    # Remove line break in rows
    new_rows = []
    #i = 0
    for row in rows:
        
        question      = row[1]
        passage_1     = row[2]
        passage_2     = row[4]
        passage_3     = row[6]
        golden_answer = row[8]

        # golden_answer = golden_answer.replace('\n', '')
        # question = question.replace('\n', '')

        new_rows.append([question, golden_answer, passage_1, passage_2, passage_3 ])

        #print(f"-----[{i}]-----\n * Question: {question}\n * Golden answer: {golden_answer}")
        #i = i + 1

    new_header = [ "question", "golden_answer", "golden_passage_1", "golden_passage_2", "golden_passage_3" ]
    
    return new_header, new_rows

# ******************************************
# load existing evaluate values from csv
def load_qa_service_metrics(csv_filepath):
        file = open(csv_filepath)
        csvreader = csv.reader(file)
        header = []
        header = next(csvreader)
        print(f"{header}")
        qa_service_metrics = []
        for row in csvreader: 
                values = [ str(row[4]) , str(row[6]) , str(row[8]) ]
                #print(f"Values:\n {values}")          
                qa_service_metrics.append(values)
        file.close()
        return qa_service_metrics

# ******************************************
# Invoke the REST API endpoint 
# of the qa service in code engine
def invoke_qa(question):
        global endpoint
        global api_url
        global username
        global password
        global output_error_log
        
        request_url = api_url + endpoint
        question_obj = {"query": ""}
        question_obj["query"] = "text:" + question

        # 1. send request
        response = requests.post(request_url, auth=(username, password), json=question_obj)
        
        print(f"Status  : {response.status_code}")
        status_code = response.status_code
        
        if (status_code != 200):
                answer_text = ""
                answer_text_len = 0
                answer_text_list = []
                # Log errors for the status_code
                message = "Response code: " + str(status_code) + "____" + "Question object: " + json.dumps(question_obj)
                print(f"Error: {message}")
                logger.error(message)
                return answer_text, answer_text_len, answer_text_list, False
        else:
                response_json = response.json()
                # 2. extract the result
                if ( response_json['results'] != False):
                        results = response_json['results']
                        result = results[0]
                        
                        # 3. verify is it an answer
                        answer = result['document_id']
                        answer_text = ""
                        answer_text_list = []

                        if (answer == verify_answer):
                                answer_text_list = result['text']
                                answer_text_len = len(answer_text_list)
                                
                                if( len(answer_text_list)>1):
                                        for answer_part in answer_text_list:
                                                #answer_text_clean = answer_part.replace('\n', '')
                                                answer_text_clean = answer_part
                                                answer_text = answer_text + ' ' + answer_text_clean
                                else:
                                        answer_text = answer_text_list[0]
                                        #answer_text = answer_text.replace('\n', '')

                                print(f"Question: {question}")
                                print(f"Answer  : {answer_text}")
                                return answer_text, answer_text_len, answer_text_list, True
                        else:
                                answer_text = ""
                                answer_text_len = 0
                                answer_text_list = []
                                return answer_text, answer_text_len, answer_text_list, False

                else:
                        answer_text = ""
                        answer_text_len = 0
                        answer_text_list = []
                        return answer_text, answer_text_len,  answer_text_list, False

# ******************************************
# Execution
def main(args):

        # Does input data exist?
        # - False: Invoke the microservice
        # - True: Use an existing csv file
        input_data_exists = False
        
        # Temp list for creating output files
        golds = [[]]

        # End experiment, when not all requests can be processed!
        end_experiment = False 
        
        # Get date to be added to the file names
        today = str(date.today())

        # Set paths for input
        output_directory = get_output_path()
        print(f"- Output dir: {output_directory}")
        input_directory = get_input_path()
        print(f"- Input dir: {input_directory}")
        input_qa_directory = get_input_qa_service_metrics_local_path()
        print(f"- Input qa dir (local): {input_qa_directory}")

        qa_metrics_run_file = input_qa_directory + "/" + output_session_id + "-Runs.csv"
        workbook_name_file = output_directory + "/"  + output_session_id + "_" + today + "_" + output_question_resp_anwser_excel
        csv_output_filepath = output_directory + "/"  + output_session_id + "_" + today + "_" +  output_question_resp_anwser

        # 1. use an input file to get the answers from the qa microserice
        if (input_data_exists == False):

                        # 1.1 load data from input file 
                        print(f"******* prepare input data from file {input_excel_filename} ********\n")    
                        excel_input_filepath = input_directory + "/" + input_excel_filename                       
                        header, rows = load_input_excel(excel_input_filepath) 
                        print(f"- Input header: {header}")                    
                        input_len=len(rows)
                        print(f"- Input len: {input_len}")
                        row = rows[0]
                        question = row[0]
                        print(f"- First question: {question}")
                        golden_answer = row[1]
                        print(f"- First golden answer: {golden_answer}")
                        
                        # 1.2. Prepare an output file for logging the execution results
                        csvfile = open(csv_output_filepath,'w',encoding='utf-8')
                        csvfile_writer = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
                        csv_line = ['count','question','answer','golden_answer','golden_passage_1','golden_passage_2','golden_passage_3','answer_passage_1','answer_passage_2','answer_passage_1']
                        stripped_line = [cell.strip() for cell in csv_line]
                        csvfile_writer.writerow(stripped_line)
                        
                        # 1.3. Prepare an output excel for logging the execution results
                        workbook = create_output_workbook(workbook_name_file)

                        # 1.4. invoke endpoint with questions and write the test results
                        i = 0
                        j = 0
                        print(f"******* invoke REST API ********\n")
                        for row in rows:
                                very_golden_answer = row[1]
                                if (len(very_golden_answer) != 0):
                                        question = row[0]
                                        golden_answer = row[1]
                                        golden_passage_1 = row[2]
                                        golden_passage_2 = row[3]
                                        golden_passage_3 = row[4]

                                        print(f"--- Request {i} ---")
                                        answer_text, answer_text_len, answer_list, verify = invoke_qa(question)
                                        
                                        # 1.4.1 Retry if the request didn't work
                                        if (verify != True):
                                                print(f"--- Retry the request {i} for {number_of_retrys } times ---")
                                                retrys = int(number_of_retrys)
                                                retrys_count = 0
                                                
                                                while (retrys_count != retrys):
                                                
                                                        print(f"Retry counter : {retrys_count} ---")
                                                        answer_text, answer_text_len, answer_list, verify = invoke_qa(question)
                                                        retrys_count = retrys_count + 1

                                                        if (verify == True):
                                                                break
                                                        else:
                                                                if (retrys_count == retrys):
                                                                        message = "END EXPERIMENT! - Retry: " + str(retrys_count) + " for request " + str(i) + " didn't work!"
                                                                        print(message)
                                                                        logger.error(message)
                                                                        end_experiment = True

                                                                        # 1.4.3 add the error to the output csv file
                                                                        csv_line = [str(i),"FAILED","FAILED","FAILED",]
                                                                        stripped_line = [cell.strip() for cell in csv_line]
                                                                        csvfile_writer.writerow(stripped_line)
                                                                        
                                                                        # 1.4.4 add the error to the output temp excel file
                                                                        # set value for cell B2=2
                                                                        j = j + 1
                                                                        workbook = openpyxl.load_workbook(workbook_name_file)
                                                                        worksheet = workbook['experiment_data']
                                                                        worksheet.cell(row=(j+1), column=1).value = "FAILED"
                                                                        worksheet.cell(row=(j+1), column=2).value = "FAILED"
                                                                        worksheet.cell(row=(j+1), column=3).value = "FAILED"
                                                                        workbook.save(workbook_name_file)
                                                                        break
                                                                else: 
                                                                        message = "Retry: " + str(retrys_count) + " for request " + str(i) + " didn't work!"
                                                                        print(message)
                                                                        logger.error(message)

                                        # 1.4.2 Request work and a anwser contains content
                                        if ((verify == True) and (len(row[1]) != 0) and ( end_experiment == False)):
                                                # 1.4.3 add the values to the output csv file
                                                csv_line = [str(i),question,answer_text,golden_answer]
                                                stripped_line = [cell.strip() for cell in csv_line]
                                                csvfile_writer.writerow(stripped_line)
                                                
                                                # 1.4.4 add the values to the output temp excel file
                                                # set value for cell B2=2
                                                j = j + 1

                                                workbook = openpyxl.load_workbook(workbook_name_file)
                                                worksheet = workbook['experiment_data']
                                                worksheet.cell(row=(j+1), column=1).value = question
                                                worksheet.cell(row=(j+1), column=2).value = answer_text
                                                worksheet.cell(row=(j+1), column=3).value = golden_answer
                                                worksheet.cell(row=(j+1), column=4).value = golden_passage_1
                                                worksheet.cell(row=(j+1), column=5).value = golden_passage_2
                                                worksheet.cell(row=(j+1), column=6).value = golden_passage_3

                                                workbook.save(workbook_name_file)

                                        else:
                                                message = "Problem in data value: " + str(i) + "____" + "Question: " + question
                                                print(f"Error: {message}")
                                                logger.error(message)
                                else:
                                        message = "Data value " + str(i) + " ____" + " 'Golden answer' is emtpy: " + row[1]
                                        print(f"Error: {message}")
                                        logger.error(message)
                                i = i + 1
                        
                        workbook.save(workbook_name_file)      
                        csvfile.close()                    
        
        if (end_experiment == False):

                  # 2. Create evalution blue result output         
                  header, rows = bleu_run(workbook_name_file)
                  header = bleu_from_list_to_dict(header)
                  responses = [row[header['response']] for row in rows]
                  golds = [[row[header['golden_anwser']]] for row in rows]
                  
                  # print(f"Response {responses}")
                  # print(f"golds {golds}")
                   
                  metric = load_metric("sacrebleu")
                  metric.add_batch(predictions=responses, references=golds)
                  sacrebleu = metric.compute()["score"]

                  metric = load_metric("rouge")
                  metric.add_batch(predictions=responses, references=golds)
                  rouge = metric.compute()["rougeL"]

                  # 3. add results from the qa service metrics

                  metrics_results = load_qa_service_metrics(qa_metrics_run_file)
                  print(f"metrics_results: {len(metrics_results)} \n")
                  workbook = openpyxl.load_workbook(workbook_name_file)
        
                  j = 1
                  for row in metrics_results:
                        worksheet = workbook['experiment_data']
                        worksheet.cell(row=(j+1), column=7).value = row[0]
                        worksheet.cell(row=(j+1), column=8).value = row[1]
                        worksheet.cell(row=(j+1), column=9).value = row[2]
                        j = j + 1
                  workbook.save(workbook_name_file)

                  # 4. Show results
                
                  print (f"******* outputs for session: {output_session_id} ********")
                  print (f"CSV   output file : {csv_output_filepath}")
                  print (f"Excel output file : {workbook_name_file}\n")
                  count = len(responses) - 1
                  print (f"******* Bleu result based on {count} responses ********")
                  print ('Bleu: ' + str(sacrebleu), 'RougeL: ' + str(rouge.mid.fmeasure))
        
        else:
                  print (f"******* Experiment failed *************")
                  print (f"******* outputs for failed session: {output_session_id} ********")
                  print (f"CSV   output file : {csv_output_filepath}")
                  print (f"Excel output file : {workbook_name_file}\n")



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    args = parser.parse_args()
    main(args)