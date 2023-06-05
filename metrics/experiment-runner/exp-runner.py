#!/usr/bin/env python

# Imports
import requests
import json
import argparse
import os
import csv
from dotenv import load_dotenv
load_dotenv()
from datasets import load_metric
import sacrebleu
import logging
import openpyxl
from openpyxl.styles import Alignment
import regex
import sys

# ******************************************
# Global variables

# ------------------------
# Environment variables

# ------------------------
# Defaults
# - Customm debug information 
if (os.environ.get("app_debug_channel") == None):
        app_debug_channel = 'True'
else:
        app_debug_channel = os.environ.get("app_debug_channel")

# ------------------------
# Defaults
# - qa_server_on_cloud
if (os.environ.get("qa_service_on_cloud") == None):
        qa_service_on_cloud = 'False'
else:
        qa_service_on_cloud = os.environ.get("qa_service_on_cloud")

# - Does input data exist?
#       - If 'False': Invoke the microservice
#            'True' : Use an existing data
if (os.environ.get("input_data_exists") == None):
        input_data_exists = 'False'
else:
        input_data_exists = os.environ.get("input_data_exists")

# - Prefix passage ID groundtruth
#       - If 'False': Invoke the microservice
#            'True' : Use an existing data
if (os.environ.get("prefix_passage_id") == None):
        prefix_passage_id = ''
else:
        prefix_passage_id = os.environ.get("prefix_passage_id")

# -----------------------
# Loaded from env file

# qa service related
endpoint = os.environ.get("endpoint")
api_url = os.environ.get("api_url")
username=os.environ.get("username")
password=os.environ.get("password")
verify_answer=os.environ.get("verify_answer")

# input for ground truth
input_excel_filename = os.environ.get("input_excel_filename")
input_folder_name = os.environ.get("input_folder_name")
input_folder_name_qa_service_metrics = os.environ.get("input_folder_name_qa_service_metrics")
container_run = os.environ.get("container_run")

# output
output_question_resp_anwser_excel = os.environ.get("output_question_resp_anwser_excel")
output_error_log = os.environ.get("output_error_log")
output_session_id = os.environ.get("output_session_id")
output_folder_name = os.environ.get("output_folder_name")
number_of_retries = os.environ.get("number_of_retries")

logger = None

# *****************************************
# Debug info

def debug_show_env_settings():

        global app_debug_channel
        # qa service
        global endpoint
        global api_url
        global username
        global password
        global verify_answer
        # input
        global input_excel_filename
        global input_folder_name
        global input_folder_name_qa_service_metrics
        global container_run
        # output
        global output_question_resp_anwser_excel
        global output_error_log 
        global output_session_id
        global output_folder_name
        global number_of_retries

        if (app_debug_channel == "True"):
                print("********** app DEBUG ***************")
                print("Experiment-runner configuration:")
                print("")
                print(f"- Endpoint: {endpoint}")
                print(f"- API URL: {api_url}")
                print(f"- retries: {number_of_retries}")
                print(f"- Username: {username}")
                print(f"- Password: {password}")
                print(f"- Verify answer: {verify_answer}\n")
                print(f"- Input Excel: {input_excel_filename}")
                print(f"- Input folder name: {input_folder_name}")
                print(f"- Input folder name qa service: {input_folder_name_qa_service_metrics}\n")
                print(f"- Output folder name: {output_folder_name}")
                print(f"- Sesssion ID output prefix: {output_session_id}")
                print(f"- Output Excel: {output_question_resp_anwser_excel}\n")
                print(f"- Error log name: {output_error_log}")
                print(f"- Container run: {container_run}")
                return True
        else:   
                return False

def debug_show_value (value):
        global app_debug_channel
        if (app_debug_channel == "True"):
                print("********** app DEBUG start ***************")
                print(value)
                print("********** app DEBUG end ***************")
                return True
        else:
                return False

# ******************************************
# get os path information
def get_input_path():
        global input_folder_name
        global container_run

        d_value = "get_input_path(): " + str(container_run)
        debug_show_value(d_value)
        
        if (container_run == "False"):
                directory = os.getcwd()
                d_value = "False: " + str(directory)
                debug_show_value(d_value)
                directory = directory + "/../" + input_folder_name
        else:
                directory = os.getcwd()
                directory = directory + "/" + input_folder_name
        return directory

def get_output_path():
        global output_folder_name
        global container_run
        if (container_run == "False"):
                directory = os.getcwd()
                new_directory = directory + "/../" + output_folder_name
        else:
                directory = os.getcwd()
                new_directory = directory + "/" + output_folder_name
        return new_directory

def get_input_qa_service_metrics_local_path():
        global input_folder_name_qa_service_metrics
        global output_folder_name
        directory = os.getcwd()
       
        if (container_run == "False"):
                directory = os.getcwd()
                d_value = "False: " + str(directory)
                debug_show_value(d_value)

                if input_folder_name_qa_service_metrics != "":
                        new_directory = directory + "/../" + output_folder_name + "/" + input_folder_name_qa_service_metrics
                else:
                        new_directory = directory + "/../" + output_folder_name
        else:
                directory = os.getcwd()
                if input_folder_name_qa_service_metrics != "":
                        new_directory = directory + "/" + output_folder_name + "/" + input_folder_name_qa_service_metrics
                else:
                        new_directory = directory + "/" + output_folder_name
             
        return new_directory

def get_input_qa_service_metrics_container_path():
        global input_folder_name_qa_service_metrics
        global output_folder_name
        global container_run
        
        if (container_run == "False"):
                directory = os.getcwd()
                new_directory = directory + "/../" + output_folder_name
        else:
                directory = os.getcwd()
                new_directory = directory + "/" + output_folder_name
        
        return new_directory

# ******************************************
# Extract data from output excel and create new data tab 
# - NA (evidence)
# - I do not have information regarding
# - Unfortunately, no relevant information is found.
#
# Defined by: 
#  Tabs: 'experiment_data' and 'experiment_filtered_data'
#  Filters: '(iv).', 'NA evidence', 'I do not have information regarding', 'Unfortunately, no relevant information is found'
def extract_unknown_response (excel_output_file):

    global logger
    workbook = openpyxl.load_workbook(excel_output_file)
    worksheet = workbook['experiment_data']
    d_value = ""
    rows = []
    not_valid_values = [ '(iv).', 'NA evidence', 'I do not have information regarding', 'Unfortunately, no relevant information is found' ]
    found_not_valid_values = []
    
    for rdx, row in enumerate(worksheet.iter_rows(values_only=True)):     
        d_value = "Load Excel output file: " + str(excel_output_file)
        debug_show_value(d_value)
        
        # End if the row contains data like this:
        # # ('None', 'None' ... )
        if not any(row):
                new_header=[]
                new_rows=[]
                message = "ERROR: Input data, please verify your output data."
                print(message)
                logger.error(message)
                return new_header, new_rows, False

        if rdx:          
            rows.append(list(row))
        else:
            header = row
            new_rows = []

    # 1. Extract data which doesn't contain wrong values
    for row in rows:
        
        for verify_value in not_valid_values:
                if verify_value in str(row[2]):
                        d_value = "Will not be added to the new result: " + str(row[2])
                        debug_show_value(d_value)
                        add_to_value = False
                        
                        question      = row[0]
                        anwer         = row[1]
                        golden_answer = row[2]
                        
                        found_not_valid_values.append([question, anwer, golden_answer, verify_value])
                        break
                else:
                        question      = row[0]
                        anwer         = row[1]
                        golden_answer = row[2]
                        add_to_value = True

        if (add_to_value == True):
                new_rows.append([question, anwer, golden_answer])

    new_header = [ "question", "anwer", "golden_answer"]

    # 2. Save the filtered values
    j = 1
    for row in new_rows:
                worksheet = workbook['experiment_filtered_data']
                d_value = "Row: \n" + str(row)
                debug_show_value(d_value)

                worksheet.cell(row=(j+1), column=1).value = str(row[0])
                worksheet.cell(row=(j+1), column=2).value = str(row[1])
                worksheet.cell(row=(j+1), column=3).value = str(row[2])
                j = j + 1
    
    worksheet = workbook['experiment_filtered_data']

    for row in worksheet.iter_rows():
        for cell in row:
                cell.alignment = Alignment(wrapText=True,vertical='top')

    # 3. Save the bad data values
    j = 1
    for row in found_not_valid_values:
                worksheet = workbook['experiment_bad_data']
                d_value = "Row: \n" + str(row)
                debug_show_value(d_value)

                worksheet.cell(row=(j+1), column=1).value = str(row[0])
                worksheet.cell(row=(j+1), column=2).value = str(row[1])
                worksheet.cell(row=(j+1), column=3).value = str(row[2])
                worksheet.cell(row=(j+1), column=4).value = str(row[3])
                j = j + 1
    
    worksheet = workbook['experiment_bad_data']

    for row in worksheet.iter_rows():
        for cell in row:
                cell.alignment = Alignment(wrapText=True,vertical='top')
    
    workbook.save(excel_output_file)

    return new_header, new_rows, found_not_valid_values, True
   
# ******************************************
# Score prepare eval data functions
def extract_prefix(input_string, prefix):
        work_string = str(input_string)
        work_prefix = str(prefix)

        if (work_string == ""):
                return_string = 'none'
                return return_string
        
        if (work_prefix == ""):
                return_string = work_string
                return return_string

        if(work_string.find(work_prefix)== -1):
                return work_string
        else:               
                return_string  = work_string.split(work_prefix,1)
                return return_string[1]

def get_score_groundtruth(excel_input_file, prefix_passage_id):
            
            wb = openpyxl.load_workbook(excel_input_file)
            ws = wb.active

            rows = []
            for rdx, row in enumerate(ws.iter_rows(values_only=True)):
                if rdx:
                        rows.append(list(row))
                else:
                        header = row
            new_rows = []

            # Extract passage data
            for row in rows:

                passage_1_id = extract_prefix(str(row[3]),prefix_passage_id)
                passage_2_id = extract_prefix(str(row[5]),prefix_passage_id)
                passage_3_id = extract_prefix(str(row[7]),prefix_passage_id)
                     
                new_rows.append([passage_1_id, passage_2_id, passage_3_id ])
                new_header = [ "passage_1_id", "passage_2_id", "passage_3_id"]
        
            return new_header, new_rows

def load_score_reranker(csv_filepath):
        d_value = "QA -service experiment file - RERANKER: " + csv_filepath
        debug_show_value(d_value)
        file = open(csv_filepath)
        csvreader = csv.reader(file)
        header = []
        header = next(csvreader)
        d_value = "QA -service experiment file Header - RERANKER: " + str(header)
        debug_show_value(d_value)

        # Reranker extract
        i = 0
        for column in header:
                # print(f"Column: {column} : {i}")
                if str(column) == "RESULT_RERANKER_PASSAGE1_ID":
                        ranker_p_1_id = i   
                if str(column) == "RESULT_RERANKER_PASSAGE2_ID":
                        ranker_p_2_id = i   
                if str(column) == "RESULT_RERANKER_PASSAGE3_ID":
                        ranker_p_3_id = i
                i = i + 1                         

        score_ranker = []
        for row in csvreader: 
                values = [ str(row[ranker_p_1_id]) , str(row[ranker_p_2_id]) , str(row[ranker_p_3_id]) ]
                # print(f"Values:\n {values}")          
                score_ranker.append(values)
        file.close()
        return score_ranker

def score_matcher(groundtruth, qa_run):

        documents_in_ground_truth = {}
        
        with open(groundtruth, 'rt') as in_f:
                
                reader = csv.DictReader(in_f, delimiter=',', quotechar='\"')
                
                for r, row in enumerate(reader):
                        for pos in ("1", "2", "3"):
                                if row['loio ' + pos] != "":
                                        documents_in_ground_truth[regex.sub(r'^loio', '', row['loio ' + pos])] = row['passage ' + pos]

                                documents_in_corpus = {}
        
        with open(qa_run) as passages_file:
                print(f'# Reading {qa_run}')
                records = json.load(passages_file)
                for rec in records:
                        id = rec['chunckid']
                        documents_in_corpus[regex.sub(r'^[0-9]\.', '', id)] = ' '.join(rec['text']), rec['title']

        num_not_found = 0
        for id in documents_in_ground_truth.keys():
                if not id in documents_in_corpus:
                        print(f'Not found: {id}')
                        num_not_found += 1

        print(f'{num_not_found} ground truth documents not found in the corpus')

        return True

def convert_groundtruth_excel_to_csv(groundtruth_xlsx,groundtruth_csv):
        wb = openpyxl.load_workbook(groundtruth_xlsx)
        sh = wb.active
        with open(groundtruth_csv, 'w', newline="" ) as f:
                c = csv.writer(f)
                for r in sh.rows:
                        c.writerow([cell.value for cell in r])
        f.close()
        return True

# ******************************************
# Bleu prepare eval data functions 

def bleu_run(input_filename, tab):
    header, rows = bleu_get_data(input_filename, tab)
    header = list(header.keys())
    return header, rows

def bleu_get_data(input_filename, tab):
    rows = list()
    header = ['question', 'response', 'golden_anwser']

    try:
        workbook = openpyxl.load_workbook(input_filename)
        ws = workbook[tab]
    except:
        print(f"Error: Could not open {input_filename}\n")
    
    for row in ws.iter_rows(values_only=True):
        rows.append(row)

    return bleu_from_list_to_dict(header), rows

def bleu_from_list_to_dict(header):
    indices = [i for i in range(0, len(header))]
    header = {k: v for k, v in zip(header, indices)}
    return header

# ******************************************
# Define logging
def create_logger():
        global output_session_id
        global output_error_log
        global logger

        print(f"{output_error_log}")

        logger = logging.getLogger(output_session_id + "-" + output_error_log)
        logger.setLevel(logging.INFO)
        path=get_output_path()
        file_handler = logging.FileHandler(path + "/" + output_session_id + "-" + output_error_log)
        file_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        return logger

# Create a "experiment_data" worksheet
def create_output_workbook (workbook_name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet("experiment_data")
        worksheet_blue = workbook.create_sheet("experiment_bleu_result")
        worksheet_experiment_filtered_data = workbook.create_sheet("experiment_filtered_data")
        worksheet_experiment_bad_data = workbook.create_sheet("experiment_bad_data")
        worksheet_performance = workbook.create_sheet("experiment_performance")

        if 'Sheet1' in  workbook.sheetnames:
                 workbook.remove( workbook['Sheet1'])
        if 'Sheet' in  workbook.sheetnames:
                 workbook.remove( workbook['Sheet'])

        # filtered data  
        worksheet_experiment_filtered_data.title = "experiment_filtered_data"
        worksheet_experiment_filtered_data['A1'] = 'question'
        worksheet_experiment_filtered_data['B1'] = 'answer'
        worksheet_experiment_filtered_data['C1'] = 'golden_anwser'
        
        # bad data
        worksheet_experiment_bad_data.title = "experiment_bad_data"
        worksheet_experiment_bad_data['A1'] = 'question'
        worksheet_experiment_bad_data['B1'] = 'answer'
        worksheet_experiment_bad_data['C1'] = 'golden_anwser'
        worksheet_experiment_bad_data['D1'] = 'bad phrase'

        # bleu_result   
        worksheet_blue.title = "experiment_bleu_result"
        worksheet_blue['A1'] = 'bleu'
        worksheet_blue['B1'] = 'RougeL'

        # performance 
        worksheet_performance.title = "experiment_performance"
        worksheet_performance['A1'] = "Timestamp start"
        worksheet_performance['B1'] = "Timestamp end"
        worksheet_performance['C1'] = "Duration"

        # data
        worksheet.title = "experiment_data"
        worksheet['A1'] = 'question'
        worksheet['B1'] = 'answer'
        worksheet['C1'] = 'golden_anwser'
        worksheet['D1'] = 'passage_1'
        worksheet['E1'] = 'passage_1_id'
        worksheet['F1'] = 'passage_2'
        worksheet['G1'] = 'passage_2_id'
        worksheet['H1'] = 'passage_3'
        worksheet['I1'] = 'passage_3_id'
        worksheet['J1'] = 'discovery_passage_1'
        worksheet['K1'] = 'discovery_passage_1_id'
        worksheet['L1'] = 'discovery_passage_2'
        worksheet['M1'] = 'discovery_passage_2_id'
        worksheet['N1'] = 'discovery_passage_3'
        worksheet['O1'] = 'discovery_passage_3_id'
        worksheet['P1'] = 'reranker_passage_1'
        worksheet['Q1'] = 'reranker_passage_1_id'
        worksheet['R1'] = 'reranker_passage_2'
        worksheet['S1'] = 'reranker_passage_2_id'
        worksheet['T1'] = 'reranker_passage_3'
        worksheet['U1'] = 'reranker_passage_3_id'
        worksheet['V1'] = 'elastic_search_passage_1'
        worksheet['W1'] = 'elastic_search_passage_1_id'
        worksheet['X1'] = 'elastic_search_passage_2'
        worksheet['Y1'] = 'elastic_search_passage_2_id'
        worksheet['Z1'] = 'elastic_search_passage_3'
        worksheet['AA1'] = 'elastic_search_passage_3_id'

        # Save the workbook as a new Excel file
        workbook.save(workbook_name)
        return workbook

# ******************************************
# load input excel
def load_input_excel(excel_input):
    global logger
    wb = openpyxl.load_workbook(excel_input)
    ws = wb.active

    d_value = "Excel input file: " + str(excel_input)
    debug_show_value(d_value)
    
    rows = []
    for rdx, row in enumerate(ws.iter_rows(values_only=True)):
        d_value = "Excel input values: " + str(row)
        debug_show_value(d_value)

        # End if the row contains data like this:
        # ('None', 'None' ... )
        if not any(row):
                new_header=[]
                new_rows=[]
                message = "ERROR: Input data, please verify your ground truth input."
                print(message)
                logger.error(message)
                return new_header, new_rows, False

        if rdx:          
            rows.append(list(row))
        else:
            header = row
    
    new_rows = []
    
    # Extract data
    for row in rows:
        
        question      = row[1]
        passage_1     = row[2]
        passage_1_id  = row[3]
        passage_2     = row[4]
        passage_2_id  = row[5]
        passage_3     = row[6]
        passage_3_id  = row[7]
        golden_answer = row[8]

        new_rows.append([question, golden_answer, passage_1, passage_1_id, passage_2, passage_2_id, passage_3, passage_3_id ])

    new_header = [ "question", "golden_answer", "passage_1", "passage_1_id", "passage_2", "passage_2_id", "passage_3", "passage_3_id"]
    
    return new_header, new_rows, True

# ******************************************
# load qa service performance from run.csv file for: 
# - TIMESTAMP_START
# - TIMESTAMP_STOP
def load_qa_service_performance(csv_filepath):
        d_value = "QA -service experiment 'run.csv' file: " + csv_filepath
        debug_show_value(d_value)
        file = open(csv_filepath)
        csvreader = csv.reader(file)
        header = []
        header = next(csvreader)
        d_value = "QA -service experiment 'run.csv' file Header: " + str(header)
        debug_show_value(d_value)

        # - TIMESTAMP_START
        # - TIMESTAMP_STOP
        i = 0
        timestamp_start = 0
        timestamp_end = 0
        for column in header:
               
                if str(column) == "TIMESTAMP_START":
                        d_value = "Column: " + str(column) + " : " + str(i)
                        debug_show_value(d_value)
                        timestamp_start = i
                if str(column) == "TIMESTAMP_END":
                        d_value = "Column: " + str(column) + " : " + str(i)
                        debug_show_value(d_value)
                        timestamp_end = i
                i = i + 1                         

        qa_service_performance = []
        total = 0
        count = 1
        for row in csvreader: 
                count = count  + 1
                d_value = "Column: " + str(row)
                debug_show_value(d_value)
                
                duration =  int(row[timestamp_end]) - int(row[timestamp_start])
                values = [ str(row[timestamp_start]) , 
                           str(row[timestamp_end]),
                           str(duration)    ]               
                
                d_value = "qa_service_performance - Values:\n " + str(values)
                debug_show_value(d_value)
                total = total + duration    
                qa_service_performance.append(values)
        
        average = int(total) / count
        file.close()
        
        return qa_service_performance, average

# ******************************************
# Add performance from the qa service metrics to output excel
def add_qa_service_performance_to_excel(qa_metrics_run_file, workbook_name_file):

        performance_results, average = load_qa_service_performance(qa_metrics_run_file)
        d_value = "Performance_results: \n" + str(len(performance_results))
        debug_show_value(d_value)
        d_value = "Performance_results: \n" + str(performance_results)
        debug_show_value(d_value)

        workbook = openpyxl.load_workbook(workbook_name_file)
        
        j = 1
        for row in performance_results:
                worksheet = workbook['experiment_performance']
                d_value = "Row: \n" + str(row)
                debug_show_value(d_value)

                worksheet.cell(row=(j+1), column=1).value = int(row[0])
                worksheet.cell(row=(j+1), column=2).value = int(row[1])
                worksheet.cell(row=(j+1), column=3).value = int(row[2])
                j = j + 1
                
        worksheet['D1']="Average"
        worksheet['D2']= average
        d_value = "Performance_results: \n" + str(worksheet['D2'])
        debug_show_value(d_value)   
        worksheet = workbook['experiment_performance']

        for row in worksheet.iter_rows():  
                for cell in row:      
                         cell.alignment = Alignment(wrapText=True,vertical='top')
               
        workbook.save(workbook_name_file)
        return True

# ******************************************
# load qa service metrics from run.csv file for: 
# - Discovery
# - Reranker
# - Elastic search
def load_qa_service_metrics(csv_filepath):
        d_value = "QA -service experiment 'run.csv' file: " + csv_filepath
        debug_show_value(d_value)
        file = open(csv_filepath)
        csvreader = csv.reader(file)
        header = []
        header = next(csvreader)
        d_value = "QA -service experiment 'run.csv' file Header: " + str(header)
        debug_show_value(d_value)

        # Reranker extract
        # Discovery extract
        # Elastic search extract
        i = 0
        for column in header:
               
                d_value = "Column: " + str(column) + " : " + str(i)
                debug_show_value(d_value)
                if str(column) == "RESULT_DISCOVERY_PASSAGE1":
                        discovery_p_1 = i
                if str(column) == "RESULT_DISCOVERY_PASSAGE2":
                        discovery_p_2 = i
                if str(column) == "RESULT_DISCOVERY_PASSAGE3":
                        discovery_p_3 = i
                if str(column) == "RESULT_DISCOVERY_PASSAGE1_ID":
                        discovery_p_1_id = i
                if str(column) == "RESULT_DISCOVERY_PASSAGE2_ID":
                        discovery_p_2_id = i
                if str(column) == "RESULT_DISCOVERY_PASSAGE2_ID":
                        discovery_p_3_id = i
                if str(column) == "RESULT_RERANKER_PASSAGE1":
                        ranker_p_1 = i
                if str(column) == "RESULT_RERANKER_PASSAGE2":
                        ranker_p_2 = i
                if str(column) == "RESULT_RERANKER_PASSAGE3":
                        ranker_p_3 = i
                if str(column) == "RESULT_RERANKER_PASSAGE1_ID":
                        ranker_p_1_id = i   
                if str(column) == "RESULT_RERANKER_PASSAGE2_ID":
                        ranker_p_2_id = i   
                if str(column) == "RESULT_RERANKER_PASSAGE3_ID":
                        ranker_p_3_id = i
                if str(column) == "RESULT_ELASTIC_PASSAGE1":
                        elastic_p_1 = i
                if str(column) == "RESULT_ELASTIC_PASSAGE2":
                        elastic_p_2 = i
                if str(column) == "RESULT_ELASTIC_PASSAGE3":
                        elastic_p_3 = i
                if str(column) == "RESULT_ELASTIC_PASSAGE1_ID":
                        elastic_p_1_id = i   
                if str(column) == "RESULT_ELASTIC_PASSAGE2_ID":
                        elastic_p_2_id = i   
                if str(column) == "RESULT_ELASTIC_PASSAGE3_ID":
                        elastic_p_3_id = i
                i = i + 1                         

        qa_service_metrics = []
        for row in csvreader: 
                values = [ str(row[discovery_p_1]) , 
                           str(row[discovery_p_1_id]) , 
                           str(row[discovery_p_2]) , 
                           str(row[discovery_p_2_id]) , 
                           str(row[discovery_p_3]), 
                           str(row[discovery_p_3_id]),
                           str(row[ranker_p_1]) , 
                           str(row[ranker_p_1_id]) , 
                           str(row[ranker_p_2]) , 
                           str(row[ranker_p_2_id]) , 
                           str(row[ranker_p_3]), 
                           str(row[ranker_p_3_id]),
                           str(row[elastic_p_1]) , 
                           str(row[elastic_p_1_id]) , 
                           str(row[elastic_p_2]) , 
                           str(row[elastic_p_2_id]) , 
                           str(row[elastic_p_3]), 
                           str(row[elastic_p_3_id]) ]
                
                d_value = "Values:\n " + str(values)
                debug_show_value(d_value)        
                qa_service_metrics.append(values)
        file.close()
        return qa_service_metrics

# ******************************************
# Add results from the qa service metrics to output excel
def add_qa_service_metrics_to_excel(qa_metrics_run_file, workbook_name_file, sacrebleu_1, str_rouge_1, sacrebleu_2, str_rouge_2 ):

        metrics_results = load_qa_service_metrics(qa_metrics_run_file)
        d_value = "Metrics_results: \n" + str(len(metrics_results))
        debug_show_value(d_value)
        d_value = "Metrics_results: \n" + str(metrics_results)
        debug_show_value(d_value)

        workbook = openpyxl.load_workbook(workbook_name_file)
        
        j = 1
        for row in metrics_results:
                worksheet = workbook['experiment_data']
                d_value = "Row: \n" + str(row)
                debug_show_value(d_value)

                worksheet.cell(row=(j+1), column=10).value = row[0]
                worksheet.cell(row=(j+1), column=11).value = row[1]
                worksheet.cell(row=(j+1), column=12).value = row[2]
                worksheet.cell(row=(j+1), column=13).value = row[3]
                worksheet.cell(row=(j+1), column=14).value = row[4]
                worksheet.cell(row=(j+1), column=15).value = row[5]
                worksheet.cell(row=(j+1), column=16).value = row[6]
                worksheet.cell(row=(j+1), column=17).value = row[7]
                worksheet.cell(row=(j+1), column=18).value = row[8]
                worksheet.cell(row=(j+1), column=19).value = row[9]
                worksheet.cell(row=(j+1), column=20).value = row[10]
                worksheet.cell(row=(j+1), column=21).value = row[11]
                worksheet.cell(row=(j+1), column=22).value = row[12]
                worksheet.cell(row=(j+1), column=23).value = row[13]
                worksheet.cell(row=(j+1), column=24).value = row[14]
                worksheet.cell(row=(j+1), column=25).value = row[15]
                worksheet.cell(row=(j+1), column=26).value = row[16]
                worksheet.cell(row=(j+1), column=27).value = row[17]
                j = j + 1
                  
                worksheet = workbook['experiment_bleu_result']
                worksheet.cell(row=(2), column=1).value = str(sacrebleu_1)
                worksheet.cell(row=(2), column=2).value = str_rouge_1
                worksheet.cell(row=(2), column=3).value = 'all data'


                worksheet = workbook['experiment_bleu_result']
                worksheet.cell(row=(3), column=1).value = str(sacrebleu_2)
                worksheet.cell(row=(3), column=2).value = str_rouge_2
                worksheet.cell(row=(3), column=3).value = 'filtered data'
        
        worksheet = workbook['experiment_data']
        for row in worksheet.iter_rows():  
                for cell in row:      
                         cell.alignment = Alignment(wrapText=True,vertical='top')
        
        worksheet = workbook['experiment_bleu_result']
        for row in worksheet.iter_rows():  
                for cell in row:      
                        cell.alignment = Alignment(wrap_text=True,vertical='top') 
        
        workbook.save(workbook_name_file)
        return True

# ******************************************
# Invoke the REST API endpoint 
# of the qa service in code engine
def invoke_qa(question):
        global endpoint
        global api_url
        global username
        global password
        global output_error_log
        global logger
        
        request_url = api_url + endpoint
        question_obj = {"query": ""}
        question_obj["query"] = "text:" + question

        # 1. send request
        response = requests.post(request_url, auth=(username, password), json=question_obj)
        
        print(f"Status  : {response.status_code}")
        status_code = response.status_code
        
        if (status_code != 200):
                answer_text = ""
                answer_text_len = 0
                answer_text_list = []
                # Log errors for the status_code
                message = "Response code: " + str(status_code) + "____" + "Question object: " + json.dumps(question_obj)
                print(f"Error: {message}")
                logger.error(message)
                return answer_text, answer_text_len, answer_text_list, False
        else:
                response_json = response.json()
                # 2. extract the result
                if ( response_json['results'] != False):
                        results = response_json['results']
                        result = results[0]
                        
                        # 3. verify is it an answer
                        answer = result['document_id']
                        answer_text = ""
                        answer_text_list = []

                        if (answer == verify_answer):
                                answer_text_list = result['text']
                                answer_text_len = len(answer_text_list)
                                
                                if( len(answer_text_list)>1):
                                        for answer_part in answer_text_list:
                                                answer_text_clean = answer_part
                                                answer_text = answer_text + ' ' + answer_text_clean
                                else:
                                        answer_text = answer_text_list[0]

                                print(f"Question: {question}")
                                print(f"Answer  : {answer_text}")
                                return answer_text, answer_text_len, answer_text_list, True
                        else:
                                answer_text = ""
                                answer_text_len = 0
                                answer_text_list = []
                                return answer_text, answer_text_len, answer_text_list, False

                else:
                        answer_text = ""
                        answer_text_len = 0
                        answer_text_list = []
                        return answer_text, answer_text_len,  answer_text_list, False

# ******************************************
# Execution
def main(args):
        logger = create_logger()
        
        # Temp list for creating output files
        golds = [[]]

        # End experiment, when not all requests can be processed!
        end_experiment = False 

        # Set paths for input and output
        output_directory = get_output_path()
        input_directory = get_input_path()
        d_value = " - Output dir: " + output_directory + "\n - Input dir: " + input_directory
        debug_show_value(d_value)

        excel_input_filepath = input_directory + "/" + input_excel_filename     
        input_name = input_excel_filename.split(".xlsx",1)
        csv_input_filepath = input_directory  + "/" + input_name[1] + ".csv"
        convert_groundtruth_excel_to_csv(excel_input_filepath,csv_input_filepath)
        
        if (container_run == "False"):
                input_qa_directory = get_input_qa_service_metrics_local_path()
                d_value = "- Input qa dir (local): " + input_qa_directory
                debug_show_value(d_value)
                qa_metrics_run_file = input_qa_directory + "/" + output_session_id + "-Runs.csv"
        else:
                input_qa_directory = get_input_qa_service_metrics_container_path()
                d_value = "- Input qa dir (container):" + input_qa_directory
                debug_show_value(d_value)
                qa_metrics_run_file = input_qa_directory + "/" + output_session_id + "-Runs.csv"
   
        workbook_name_file = output_directory + "/"  + output_session_id + "-" + output_question_resp_anwser_excel

        # 1. use an input file to get the answers from the qa microserice
        if (input_data_exists == "False"):

                        # 1.1 Load data from input file 
                        d_value = "******* prepare input data from file " + input_excel_filename + " ********\n"
                        debug_show_value(d_value)                      
                        header, rows, check = load_input_excel(excel_input_filepath) 
                        
                        # 1.1.1 Exit when the problems with the input data
                        if (check == False):
                                end_experiment == True
                                sys.exit("Problem with the input data the automation ends here.")
                        
                        end_experiment == False
                        d_value = "- Input header: " + str(header)
                        debug_show_value(d_value)                  
                        input_len=len(rows)
                        
                        d_value = "- Input len: " + str(input_len)
                        debug_show_value(d_value)
                        row = rows[0]
                        question = row[0]
                        
                        d_value = "First question: " + question
                        debug_show_value(d_value)
                        golden_answer = row[1]
                        d_value = "- First golden answer: " + golden_answer
                        debug_show_value(d_value)
                               
                        # 1.2. Prepare an output excel for logging the execution results
                        workbook = create_output_workbook(workbook_name_file)

                        # 1.3. invoke endpoint with questions and write the test results
                        i = 0
                        j = 0
                        print(f"******* invoke REST API ********\n")
                        for row in rows:
                                very_golden_answer = row[1]
                                if (end_experiment == True):
                                        break
                                d_value = "Row:" + str(row) + "Length:" + str(len(row))
                                debug_show_value(d_value)
 
                                if (len(very_golden_answer) != 0):
                                        question      = row[0]
                                        golden_answer = row[1]
                                        passage_1     = row[2]
                                        passage_1_id  = row[3]
                                        passage_2     = row[4]
                                        passage_2_id  = row[5]
                                        passage_3     = row[6]
                                        passage_3_id  = row[7]

                                        print(f"--- Request {i} ---")
                                        answer_text, answer_text_len, answer_list, verify = invoke_qa(question)
                                        
                                        # 1.4.1 Retry if the request didn't work
                                        if (verify != True):
                                                print(f"--- Retry the request {i} for {number_of_retries } times ---")
                                                retries = int(number_of_retries)
                                                retries_count = 0
                                                
                                                while (retries_count != retries):
                                                
                                                        print(f"Retry counter : {retries_count} ---")
                                                        answer_text, answer_text_len, answer_list, verify = invoke_qa(question)
                                                        retries_count = retries_count + 1

                                                        if (verify == True):
                                                                break
                                                        else:
                                                                if (retries_count == retries):
                                                                        message = "END EXPERIMENT! - Retry: " + str(retries_count) + " for request " + str(i) + " didn't work!"
                                                                        print(message)
                                                                        logger.error(message)
                                                                        end_experiment = True
                                                                        
                                                                        # 1.4.2 add the error to the output temp excel file
                                                                        # set value for cell B2=2
                                                                        j = j + 1
                                                                        workbook = openpyxl.load_workbook(workbook_name_file)
                                                                        worksheet = workbook['experiment_data']
                                                                        worksheet.cell(row=(j+1), column=1).value = "FAILED"
                                                                        worksheet.cell(row=(j+1), column=2).value = "FAILED"
                                                                        worksheet.cell(row=(j+1), column=3).value = "FAILED"
                                                                        workbook.save(workbook_name_file)
                                                                        break
                                                                else: 
                                                                        message = "Retry: " + str(retries_count) + " for request " + str(i) + " didn't work!"
                                                                        print(message)
                                                                        logger.error(message)

                                        # 1.4.3 Request work and a anwser contains content
                                        if ((verify == True) and (len(row[1]) != 0) and ( end_experiment == False)):

                                                # 1.4.4 add the values to the output temp excel file
                                                # set value for cell B2=2
                                                j = j + 1

                                                workbook = openpyxl.load_workbook(workbook_name_file)
                                                worksheet = workbook['experiment_data']
                                                worksheet.cell(row=(j+1), column=1).value = question
                                                worksheet.cell(row=(j+1), column=2).value = answer_text
                                                worksheet.cell(row=(j+1), column=3).value = golden_answer
                                                worksheet.cell(row=(j+1), column=4).value = passage_1
                                                worksheet.cell(row=(j+1), column=5).value = passage_1_id
                                                worksheet.cell(row=(j+1), column=6).value = passage_2
                                                worksheet.cell(row=(j+1), column=7).value = passage_2_id
                                                worksheet.cell(row=(j+1), column=8).value = passage_3
                                                worksheet.cell(row=(j+1), column=9).value = passage_3_id

                                                workbook.save(workbook_name_file)

                                        else:
                                                message = "Problem in data value: " + str(i) + "____" + "Question: " + question
                                                print(f"Error: {message}")
                                                logger.error(message)
                                else:
                                        message = "Data value " + str(i) + " ____" + " 'Golden answer' is emtpy: " + row[1]
                                        print(f"Error: {message}")
                                        logger.error(message)
                                i = i + 1
                        
                        workbook.save(workbook_name_file)                       
        
        if (end_experiment == False):

                  #header, ground_truth_rows = get_score_groundtruth(excel_input_filepath,prefix_passage_id)
                  #score_ranker = load_score_reranker(qa_metrics_run_file)
                  #score_matcher(csv_input_filepath, qa_metrics_run_file)

                  # 2. Create experiment-runner blue result output  
                  # 2.1 First calc 
                  # - with experiment_data  
                  d_value = "********** Bleu 1 start ***********"
                  debug_show_value(d_value)
       
                  header, rows = bleu_run(workbook_name_file,'experiment_data')
                  header = bleu_from_list_to_dict(header)
                  
                  debug_show_value(d_value)
                  d_value = "Header for bleu: " + str(header)
                  debug_show_value(d_value)
                  d_value = "Rows for bleu: " + str(rows)
                  debug_show_value(d_value)

                  responses_1 = [row[header['response']] for row in rows]
                  d_value = "Responses: " + str(responses_1)
                  debug_show_value(d_value)
                  golds_1 = [[row[header['golden_anwser']]] for row in rows]
                   
                  metric = load_metric("sacrebleu")
                  metric.add_batch(predictions=responses_1, references=golds_1)
                  sacrebleu_1 = metric.compute()["score"]

                  metric = load_metric("rouge")
                  metric.add_batch(predictions=responses_1, references=golds_1)
                  rouge_1 = metric.compute()["rougeL"]
                  
                  # Second calc: 
                  # - with experiment_filtered_data
                  if (qa_service_on_cloud == 'False'):
                        new_header, new_rows, found_not_valid_values, extract_result = extract_unknown_response(workbook_name_file)

                  d_value = "********** Bleu 2 start ***********"
                  debug_show_value(d_value)

                  header, rows = bleu_run(workbook_name_file,'experiment_filtered_data')
                  header = bleu_from_list_to_dict(header)
                  d_value = "Header for bleu: " + str(header)
                  debug_show_value(d_value)
                  d_value = "Rows for bleu: " + str(rows)
                  debug_show_value(d_value)

                  responses_2 = [row[header['response']] for row in rows]
                  d_value = "Responses: " + str(responses_2)
                  debug_show_value(d_value)
                  golds_2 = [[row[header['golden_anwser']]] for row in rows]
                   
                  metric = load_metric("sacrebleu")
                  metric.add_batch(predictions=responses_2, references=golds_2)
                  sacrebleu_2 = metric.compute()["score"]

                  metric = load_metric("rouge")
                  metric.add_batch(predictions=responses_2, references=golds_2)
                  rouge_2 = metric.compute()["rougeL"]

                  # 3. Add results from the qa service metrics to output excel
                  d_value = "Qa_service_on_cloud: \n" + str(qa_service_on_cloud)
                  debug_show_value(d_value)
                  if (qa_service_on_cloud == 'False'):
                        add_qa_service_metrics_to_excel(qa_metrics_run_file, 
                                                        workbook_name_file, 
                                                        str(sacrebleu_1), 
                                                        str(rouge_1.mid.fmeasure),
                                                        str(sacrebleu_2), 
                                                        str(rouge_2.mid.fmeasure))
                        add_qa_service_performance_to_excel(qa_metrics_run_file, workbook_name_file)

                  # 4. Show results
                
                  print (f"******* outputs for session: {output_session_id} ********")
                  print (f"Excel output file : {workbook_name_file}\n")
                  count_1 = len(responses_1) - 1
                  count_2 = len(responses_2) - 1
                  print (f"******* Bleu result based on {count_1} responses ********")
                  print ('Bleu: ' + str(sacrebleu_1), 'RougeL: ' + str(rouge_1.mid.fmeasure))
                  print (f"******* Filtered bleu result based on {count_2} responses ********")
                  print ('Bleu: ' + str(sacrebleu_2), 'RougeL: ' + str(rouge_2.mid.fmeasure))
        
        else:
                  print (f"******* Experiment failed *************")
                  print (f"******* outputs for failed session: {output_session_id} ********")
                  print (f"Excel output file : {workbook_name_file}\n")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    args = parser.parse_args()
    main(args)